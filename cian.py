#!/usr/bin/env python3
import statistics
from fake_useragent import UserAgent
import requests
import hashlib
import openpyxl
import sys, os


def XLSXDictReader(f):
    book = openpyxl.reader.excel.load_workbook(f)
    sheet = book.active
    rows = sheet.max_row
    cols = sheet.max_column
    headers = dict((i, sheet.cell(row=1, column=i).value) for i in range(1, cols))

    def item(i, j):
        return (sheet.cell(row=1, column=j).value, sheet.cell(row=i, column=j).value)

    return (dict(item(i, j) for j in range(1, cols + 1)) for i in range(2, rows + 1))


def get_data_from_cian(url):
    url_to_xls = 'https://www.cian.ru/export/xls/offers/?' + url.split('?')[1]
    ua = UserAgent()
    headers = {'User-Agent': ua.random}
    r = requests.get(url_to_xls, stream=True, headers=headers)
    if r.status_code == 200:
        local_filename = hashlib.md5(url.encode('utf-8')).hexdigest() + '.xlsx'
        with open(local_filename, 'wb') as f:
            for chunk in r.iter_content(chunk_size=1024):
                if chunk:  # filter out keep-alive new chunks
                    f.write(chunk)
        flats_list = list(XLSXDictReader(local_filename))
        flat_dict = {}
        flat_dict_whole = {}
        for flat in flats_list:
            numbers_of_room = int(flat['Количество комнат'].split(',')[0])
            if 'руб' not in flat['Цена']:
                continue
            cost = float(flat['Цена'].split(' руб')[0])
            area = float(flat['Площадь, м2'].split('/')[0])
            try:
                flat_dict[numbers_of_room].append(cost / area)
                flat_dict_whole[numbers_of_room].append(cost)
            except KeyError:
                flat_dict[numbers_of_room] = [cost / area]
                flat_dict_whole[numbers_of_room] = [cost]
        os.remove(local_filename)
        return flat_dict, flat_dict_whole
    else:
        print("status_code is: {}".format(r.status_code))


if __name__ == "__main__":
    url = sys.argv[1]
    flat_dict, flat_dict_whole = get_data_from_cian(url)
    rooms = sorted(list(flat_dict))
    for num_of_rooms in rooms:
        print("=" * 80)
        print("number of rooms: {}".format(num_of_rooms))
        print("number of flats: {}".format(len(flat_dict[num_of_rooms])))
        print("harmonic_mean for 1^2m: {:.2f}".format(statistics.harmonic_mean(flat_dict[num_of_rooms])))
        print("median for 1^2m: {:.2f}".format(statistics.median(flat_dict[num_of_rooms])))
        print("full price harmonic: {:.2f}".format(statistics.harmonic_mean(flat_dict_whole[num_of_rooms])))
        print("full price median: {:.2f}".format(statistics.median(flat_dict_whole[num_of_rooms])))
